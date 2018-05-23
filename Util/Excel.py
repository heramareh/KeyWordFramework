#encoding=utf-8
import re
from copy import deepcopy
from openpyxl import *
from openpyxl.styles import Border, Side, Font
import time
import os
from ProjectVar import var
from ProjectVar.var import test_data_excel_path
from Util.FormatTime import date_time_chinese


class ParseExcel(object):
    u"""Excel操作类"""
    def __init__(self, excel_file_path, sheet_name=None, new_file_path=None):
        self.excel_file_path = excel_file_path
        self.new_file_path = new_file_path
        self.workbook = load_workbook(self.excel_file_path)
        if sheet_name:
            self.worksheet = self.workbook.get_sheet_by_name(sheet_name)
        else:
            self.worksheet = self.workbook.active
        self.font = Font(color=None)
        self.colorDict = {"red":'FFFF3030', "green":'FF008B00'}

    def get_sheet_by_name(self, sheet_name):
        # return worksheet
        self.worksheet = self.workbook.get_sheet_by_name(sheet_name)
        return self.worksheet

    def get_all_sheet_names(self):
        return self.workbook.get_sheet_names()

    def get_sheet_by_index(self, sheet_index):
        # return worksheet
        self.worksheet = self.workbook.worksheets[sheet_index]
        return self.worksheet

    def get_sheet_name(self):
        # return sheet_name
        return self.worksheet.title

    def get_max_row_number(self):
        # return max_row_number
        return self.worksheet.max_row

    def get_min_row_number(self):
        # return min_row_number
        return self.worksheet.min_row

    def get_max_col_number(self):
        # return max_col_number
        return self.worksheet.max_column

    def get_min_col_number(self):
        # return min_col_number
        return self.worksheet.min_column

    def get_all_rows(self):
        #return rows
        return self.worksheet.rows

    def get_all_cols(self):
        #return cols
        return self.worksheet.columns

    def get_single_row(self, row_no):
        #return row
        return self.get_all_rows()[row_no - 1]

    def get_single_col(self, col_no):
        #return col
        return self.get_all_cols()[col_no - 1]

    def get_cell(self,row_no,col_no):
        #return cell
        return self.worksheet.cell(row=row_no, column=col_no)

    def get_cell_content(self,row_no,col_no):
        #return cell.value
        return self.get_cell(row_no, col_no).value

    def get_all_rows_content(self):
        # return datas
        datas = []
        for row in self.get_all_rows():
            data = []
            for cell in row:
                value = cell.value
                if value == None:
                    data.append(u'')
                else:
                    data.append(value)
            datas.append(data)
        return datas

    def write_cell_content(self,row_no,col_no,content,font=None):
        #return none
        self.get_cell(row_no, col_no).value = content
        self.save_file(self.new_file_path)

    def write_cell_current_time(self,row_no,col_no):
        #return none
        self.get_cell(row_no, col_no).value = date_time_chinese()
        self.save_file(self.new_file_path)

    def save_file(self, excel_file_path=None):
        if excel_file_path:
            self.workbook.save(excel_file_path)
        else:
            self.workbook.save(self.excel_file_path)

    def get_data_dict(self):
        u"""文件第一行为key，从第二行往后为value"""
        datas = []
        for row in xrange(2, self.get_max_row_number()+1):
            data = {}
            for col in xrange(1, self.get_max_col_number()+1):
                value = self.get_cell_content(row,col)
                if value == None:
                    data[self.get_cell_content(1, col)] = ''
                else:
                    data[self.get_cell_content(1, col)] = value
            datas.append(data)
        return datas

    def get_title(self):
        u"""获取第一行标题"""
        data = []
        for col in xrange(1, self.get_max_col_number()+1):
            data.append(self.get_cell_content(1, col))
        return data

    def get_values_and_actions_dict(self):
        u"""获取数据和操作以字典返回"""
        result = {}
        values, actions = self.get_values_and_actions_arr()
        result['values_title'] = values[0]
        result['actions_title'] = actions[0]
        result['values'] = self.exchange_arrs_to_dicts(values)
        result['actions'] = self.exchange_arrs_to_dicts(actions)
        return result

    def get_values_and_actions_arr(self):
        u"""获取数据和操作以数组返回"""
        values, actions = [], []
        datas = self.get_all_rows_content()
        for id, data in enumerate(datas):
            if list(set(data)) == ['']:
                actions, values = datas[:id], datas[id + 1:]
                break
        return values, actions

    def exchange_arrs_to_dicts(self, arrs):
        u"""将一个二维数组转换成字典数组"""
        results = []
        keys = arrs[0]
        for arr in arrs[1:]:
            result = {}
            for id, key in enumerate(keys):
                if key:
                    result[key] = arr[id]
            results.append(result)
        return results

    def get_replace_params_action(self, actions, value):
        u"""获取替换参数后的action"""
        result = deepcopy(actions)
        for id, action in enumerate(result):
            action_values = re.findall("\$\((.*)\)", action[var.actionValue])
            is_run = re.match("\$\((.*)\)", action[var.isRun])
            locator_expressions = re.findall("\$\((.*)\)", action[var.locatorExpression])
            for action_value in action_values:
                result[id][var.actionValue] = result[id][var.actionValue].replace("$("+action_value+")", value[action_value])
            if is_run:
                result[id][var.isRun] = value[is_run.group(1)]
            for locator_expression in locator_expressions:
                result[id][var.locatorExpression] = result[id][var.locatorExpression].replace("$("+locator_expression+")", value[locator_expression])
        return result

    def get_replace_params_actions(self):
        u"""获取替换参数后的actions"""
        dicts = self.get_values_and_actions_dict()
        actions = dicts['actions']
        values = dicts['values']
        results = []
        for value in values:
            results.append(self.get_replace_params_action(actions, value))
        return results

    def get_commants(self):
        u"""获取拼接完成后的所有操作，以数组返回"""
        commands = []
        for action in self.get_replace_params_actions():
            command = []
            for id, data in enumerate(action):
                if data[var.isRun] not in [u'off', u'否', u'False', u'0']:
                    action_list = [data[var.locatorMethod], data[var.locatorExpression], data[var.actionValue]]
                    s = ','.join(action_list).strip(',').replace(',', '","')
                    params = '"' + s + '"' if s else ''
                    command_line = data[var.actionName] + "(" + params + ")"
                    # print command_line
                    command.append(command_line)
            commands.append(command)
        return commands

if __name__ == '__main__':
    # pe = ParseExcel("e:\\test.xlsx", new_file_path="e:\\tset3.xlsx")
    # print pe.get_max_row_number()
    # print pe.get_min_row_number()
    # print pe.get_max_col_number()
    # print pe.get_min_col_number()
    # print pe.get_all_rows()
    # print pe.get_all_cols()
    # print pe.get_cell(1, 1)
    # print pe.get_cell_content(1, 1)
    # for cell in pe.get_single_row(1):
    #     print cell.value,
    # print
    # print pe.get_single_col(1)
    # pe.write_cell_content(3, 6, "哈哈")
    # pe.write_cell_current_time(3, 5)
    pe = ParseExcel(test_data_excel_path, u"联系人")
    for commant in pe.get_commants():
        print commant